from utils import calculate_ot_ut, close_excel_if_open, get_empl_working_hour, load_excel, open_excel
from openpyxl.worksheet.worksheet import Worksheet
import logging
from datetime import datetime
import os 
import openpyxl.styles as style
from openpyxl.utils import get_column_letter


# Setup logging
logging.basicConfig(
    level=logging.DEBUG,  # Change to INFO for less detail
    format="%(asctime)s [%(levelname)s] %(message)s"
)


def write_to_excel(data, excile_path):
    logging.info(f"Opening workbook: {excile_path}")
    wb = load_excel(excile_path)
    
    if wb == None:
        new_path = input("Provide Correct excel path: ")
        if new_path.startswith("\"") and new_path.endswith("\""):
            new_path = new_path[1:len(new_path)-1]
            
        write_to_excel(data, new_path)
        return
    
    ws = wb.active
    close_excel_if_open(excile_path)

    # Get month from Excel sheet (row 3, merged col FGH)
    sheet_month = ws["F3"].value.strip().upper() if ws["F3"].value else None
    logging.info(f"Sheet month: {sheet_month}")

    for entry in data:
        emp_id = entry["employee_code"]
        emp_name = entry["employee_fname"]
        in_time = entry["In_time"]
        out_time = entry["Out_time"]
        att_date = entry["Att_month"]  

        logging.debug(f"Processing employee {emp_id} - {emp_name} | Date: {att_date}")
        
        # Month check
        month_str = att_date.strftime("%b").upper()
        if month_str != sheet_month:
            logging.error(f"Skipping {emp_id}: Month {month_str} does not match sheet month {sheet_month}")
            return


        if not in_time or not out_time:
            logging.warning(f"Skipping {emp_id} - {emp_name}: Missing in/out time")
            continue

        # Convert to datetime if string
        if isinstance(in_time, str):
            in_time = datetime.fromisoformat(in_time)
        if isinstance(out_time, str):
            out_time = datetime.fromisoformat(out_time)
        if isinstance(att_date, str):
            att_date = datetime.fromisoformat(att_date)


        # Find employee row block (3 rows per employee)
        emp_row = None
        for row in range(8, ws.max_row, 3):
            emp_code = ws.cell(row=row, column=2).value  # Col B
            if str(emp_code).strip() == str(emp_id).strip():
                emp_row = row
                break

        if not emp_row:
            logging.warning(f"Employee {emp_id} not found in sheet")
            continue
        else:
            logging.debug(f"Employee row found: {emp_row}")


        # Find column for date
        date_col = None
        for col in range(6, ws.max_column + 1):  # Starting F=6
            in_time_date = int(in_time.strftime("%d"));

            if (col-5) == in_time_date:
                date_col = col
                break
        
        if not date_col:
            logging.warning(f"Date {in_time.date()} not found for {emp_id}")
            continue
        else:
            logging.debug(f"Date column found: {date_col}")
            
                
        # Getting the day of attendance date
        mention_day_in_excel = ws.cell(row=7, column=date_col).value
        print("{DEBUG} mention_day_in_excel  -> ", mention_day_in_excel)
        actull_day = att_date.strftime("%a")
        
        if mention_day_in_excel.lower() != actull_day.lower():
            logging.warning(f"Please Correct the Day in excel for this date - {att_date.strftime("%B %d, %Y")}")

        # ===== Apply Attendance Rules =====
        att_mark = ""
        overtime_hours = ""
        working_hour = 8.5
        sunday_duty = False
        
        # Checking users working hour from json
        working_hour, sunday_duty = get_empl_working_hour(emp_code)
    
        if in_time.strftime("%H:%M") == "00:00" and out_time.strftime("%H:%M") == "00:00":
            att_mark = "A" if actull_day.lower() != "sun" and sunday_duty == False else "P"
        elif in_time.strftime("%H:%M") != "00:00" and out_time.strftime("%H:%M") == "00:00":
            att_mark = "MIS"
        elif in_time.strftime("%H:%M") == out_time.strftime("%H:%M"):
            att_mark = "MIS"
        else:
            att_mark = "P"
            overtime_hours = calculate_ot_ut(
                in_time, 
                out_time, 
                actull_day, 
                sunday_duty=sunday_duty, 
                standard_hours=working_hour
            )
            
        # ===== Write into Excel =====
        logging.info(f"Writing for {emp_id} - {emp_name} | {date_col}:{emp_row} | Date: {in_time.date()} | Mark: {att_mark} | OT: {overtime_hours}")
        ws.cell(row=emp_row, column=date_col, value=att_mark)     
        ws.cell(row=emp_row + 1, column=date_col, value=overtime_hours)  

    wb.save(excile_path)
    logging.info(f"Workbook saved: {excile_path}")
    open_excel(excile_path)
    


    
# --- Predefine border styles ---
thin = style.Side(style="thin", color="000000")
thick = style.Side(style="thick", color="000000")
thin_border = style.Border(top=thin, left=thin, right=thin, bottom=thin)
thick_border = style.Border(top=thick, left=thick, right=thick, bottom=thick)
    

def write_report(ws: Worksheet, title: str, data_columns: list[str], data: list[dict]):
    """Writes a report section to the given worksheet with borders, title, and table formatting."""
    
    # --- Find next empty row ---
    row_to_write = ws.max_row + 2
    # if row_to_write != 1 and ws.cell(row=row_to_write, column=1).value not in (None, "", " "):
    #     row_to_write += 2  # leave a gap
    
    no_of_col = len(data_columns)

    # --- Title Row ---
    ws.merge_cells(start_row=row_to_write, start_column=1, end_row=row_to_write, end_column=no_of_col)
    title_cell = ws.cell(row=row_to_write, column=1, value=title)
    title_cell.font = style.Font(bold=True, size=16)
    title_cell.alignment = style.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_to_write].height = 28
    
    # Apply thick border around merged title row
    for col in range(1, no_of_col + 1):
        ws.cell(row=row_to_write, column=col).border = style.Border(top=thick, left=thick, right=thick, bottom=thin)

    # --- Column Headers ---
    row_to_write += 1
    for col_idx, col_name in enumerate(data_columns, start=1):
        cell = ws.cell(row=row_to_write, column=col_idx, value=col_name)
        cell.font = style.Font(bold=True)
        cell.alignment = style.Alignment(horizontal="center", vertical="center")

        # Thicker outer borders
        left = thick if col_idx == 1 else thin
        right = thick if col_idx == no_of_col else thin
        cell.border = style.Border(left=left, right=right, bottom=thin)

    # --- Data Rows ---
    for i, entry in enumerate(data):
        row_to_write += 1 
        is_last_row = (i == len(data) - 1)
        for col_idx, key in enumerate(data_columns, start=1):
            cell = ws.cell(row=row_to_write, column=col_idx, value=entry[key])

            # Alignment
            if key in ("Shift Hours", "Reason", "Working Hour"):
                cell.alignment = style.Alignment(horizontal="center", vertical="center")

            # Border logic
            left = thick if col_idx == 1 else thin
            right = thick if col_idx == no_of_col else thin
            top = thin
            bottom = thick if is_last_row else thin

            cell.border = style.Border(top=top, left=left, right=right, bottom=bottom)
    
    
    
    # --- Auto Adjust Column Widths ---
    for col_idx, col_name in enumerate(data_columns, start=1):
        column_letter = get_column_letter(col_idx)  
        max_length = len(str(col_name))

        # find max length among all data values in that column
        for entry in data:
            value = str(entry.get(col_name, ""))
            if len(value) > max_length:
                max_length = len(value)

        # set width slightly larger than max content length
        ws.column_dimensions[column_letter].width = max_length + 3
            
