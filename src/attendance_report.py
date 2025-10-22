
from datetime import datetime
import os
import sys
from openpyxl import Workbook
import pandas as pd
from logger import logger 
from fetcher import fetching_report
from utils import open_excel, get_valid_date
from writer import write_report


# ===== User input =======
date = get_valid_date()




def preparing_excel(file_name: str):
    wb = None
    try: 
        if os.path.isfile(file_name):
            os.remove(file_name)
        wb = Workbook()        
        wb.save(file_name)
        return wb
    except Exception as e:
        logger.error(f"Unable to open or create excel\nError: {e}")
        return None
 
 
def get_or_create_worksheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    return ws




if __name__ == "__main__":
    try: 
    
        # Fetching and analyzing attendance data
        daily_report_data = fetching_report(date) 
        reports = list(daily_report_data.keys())
        date = datetime.fromisoformat(f"{daily_report_data["report_date"]}")

        file_name = f"Daily Report {date.strftime("%d-%m-%Y")}.xlsx"
        
        # preparting a excel to write data
        wb = preparing_excel(file_name)
        if wb == None:
            input("Script exit error with code 1")
            sys.exit()
        
        # Writing each report in seprate excel sheet
        for report_type in reports:
            if report_type == 'report_date':  # skipping for date
                continue 
            
            # Finding report data for each report 
            report_data = daily_report_data[report_type]
            dataColumns = list(report_data[0].keys())
            
            logger.info(f"Writing report for {report_type}")
            ws = wb.active
            
            # setting up sheet view and style
            ws.sheet_view.showGridLines = False
            write_report(
                ws=ws,
            title= f"{report_type.replace("_", " ").capitalize()}", 
            data_columns= dataColumns, 
            data=report_data
            )
        
        wb.save(file_name)
        logger.info(f"Report saved successfully at")
        open_excel(file_name)
        input("Report is ready (Press Enter to close): ")
    except Exception as e:
        logger.error(f"Report generation fail: {e}")

        